"""
Copyright (c) 2013 Gabriel Kind <gkind@hs-mittweida.de>
Hochschule Mittweida, University of Applied Sciences

Released under the MIT license
"""
from collections import OrderedDict

from Bio import SeqIO
from django.core.exceptions import ValidationError

import cyano.models as cmodels
from cyano.helpers import slugify
from bioparser import BioParser
from django.db.transaction import commit_on_success


class _NoOverwriteDict(dict):
    # overwriting still possible with update
    def __setitem__(self, key, value):
        if key in self:
            raise ValueError("Key {} already in use".format(key))
        super(_NoOverwriteDict, self).__setitem__(key, value)


class EasyProt(BioParser):
    def __init__(self, wid, user, reason):
        super(EasyProt, self).__init__(wid, user, reason)

        self.export_parameters = _NoOverwriteDict()
        self.job_params = _NoOverwriteDict()
        self.target_peptides = []
        self.target_peptides_header = []
        self.decoy_peptides = []
        self.decoy_peptides_header = []
        self.protein_details = []
        self.protein_details_header = []
        self.protein_summary = []
        self.protein_summary_header = []
        self.mascat_quant_details_header = OrderedDict()
        self.mascat_quant_details = {}
        self.libra_quant_details_header = OrderedDict()
        self.libra_quant_details = {}
        # PTM: Post Translational Modifications

    def parse(self, handle):
        if hasattr(self, "notify_progress"):
            self.notify_progress(current=0, total=1, message="Parsing EasyProt file...")

        from openpyxl.reader.excel import load_workbook
        workbook = load_workbook(handle)

        sheet_export_parameters = workbook.get_sheet_by_name("Export Parameters")
        if sheet_export_parameters is None:
            raise ValidationError("Sheet Export Parameters missing")

        sheet_jobs_params = workbook.get_sheet_by_name("Jobs Params")
        if sheet_jobs_params is None:
            raise ValidationError("Sheet Jobs Params missing")

        self._parse_export_parameters(sheet_export_parameters)
        self._parse_jobs_params(sheet_jobs_params)

        sheet_target_peptides = workbook.get_sheet_by_name("Target Peptides")
        if sheet_target_peptides is not None:
            self._parse_target_peptides(sheet_target_peptides)

        sheet_decoy_peptides = workbook.get_sheet_by_name("Decoy Peptides")
        if sheet_decoy_peptides is not None:
            self._parse_decoy_peptides(sheet_decoy_peptides)

        sheet_protein_details = workbook.get_sheet_by_name("Protein Details")
        if sheet_protein_details is not None:
            self._parse_protein_details(sheet_protein_details)

        sheet_protein_summary = workbook.get_sheet_by_name("Protein Summary")
        if sheet_protein_summary is not None:
            self._parse_protein_summary(sheet_protein_summary)

        sheet_quant_stats_info = workbook.get_sheet_by_name("Quant Stats Info")
        if sheet_quant_stats_info is not None:
            self._parse_quant_stats_info(sheet_quant_stats_info)

        sheet_mascat_quant_details = workbook.get_sheet_by_name("Mascat Quant Details")
        if sheet_mascat_quant_details is not None:
            self._parse_mascat_quant_details(sheet_mascat_quant_details)

        sheet_libra_quant_details = workbook.get_sheet_by_name("Libra Quant Details")
        if sheet_libra_quant_details is not None:
            self._parse_libra_quant_details(sheet_libra_quant_details)

        sheet_protein_expression = workbook.get_sheet_by_name("Protein Expression 95%")
        if sheet_protein_details is not None:
            self._parse_protein_expression(sheet_protein_expression)

    @staticmethod
    def _add_to_dict(dictionary, key, value):
        if key in dictionary:
            raise ValidationError("Key {} already in use".format(key))
        dictionary[key] = value

    def _parse_export_parameters(self, sheet_export_parameters):
        """
        :type sheet_export_parameters: openpyxl.worksheet.Worksheet
        """
        row = sheet_export_parameters.get_highest_row()
        parse_jobs = False
        parse_isotopic = False
        for i in range(row):
            typ = sheet_export_parameters.cell(row=i, column=0).value
            value = sheet_export_parameters.cell(row=i, column=1).value

            if typ is not None:
                typ = typ.lower()

            if parse_jobs:
                if typ.startswith("job"):
                    value2 = sheet_export_parameters.cell(row=i, column=2).value
                    self.export_parameters["jobs"].append([value, value2])
                    continue
                else:
                    parse_jobs = False
            elif parse_isotopic:
                if typ:
                    parse_isotopic = False
                else:
                    self.export_parameters["isotopic purity correction"].append(value)
                    continue

            if typ is None:
                continue

            if typ == "easyprot version":
                typ = "version"
            elif typ == "#jobs":
                parse_jobs = True
                self.export_parameters["jobs"] = []
                continue
            elif typ.startswith("job"):
                raise ValidationError("Job outside of #Jobs list")
            elif typ == "isotopic purity correction":
                parse_isotopic = True
                self.export_parameters["isotopic purity correction"] = []
                self.export_parameters["isotopic purity correction"].append(value)
                continue

            self.export_parameters[typ] = value

    def _parse_jobs_params(self, sheet_jobs_params):
        """
        :type sheet_jobs_params: openpyxl.worksheet.Worksheet
        """
        # TODO: Testen of Job Title und ID bekannt sind in Export params
        self.sheet_export_parameters["jobs"]
        any(["FIXME job title","FIXME job id"] == x for x in self.sheet_export_parameters["jobs"])
        row = sheet_jobs_params.get_highest_row()
        parse_jobs = False
        for i in range(row):
            typ = sheet_jobs_params.cell(row=i, column=0).value
            value = sheet_jobs_params.cell(row=i, column=1).value

            typ = typ.lower()

            if parse_jobs:
                if typ.startswith("job"):
                    value2 = sheet_jobs_params.cell(row=i, column=2).value
                    self.export_parameters["jobs"].append([value, value2])
                    continue
                else:
                    parse_jobs = False

            if typ == "easyprot version":
                typ = "version"
            elif typ == "#jobs":
                parse_jobs = True
                self.export_parameters["jobs"] = []
                continue
            elif typ.startswith("job"):
                raise ValidationError("Job outside of #Jobs list")

            self.export_parameters[typ] = value

    def _parse_target_peptides(self, sheet_target_peptides):
        self._read_table(sheet_target_peptides, self.target_peptides_header, self.target_peptides)

    def _parse_decoy_peptides(self, sheet_decoy_peptides):
        self._read_table(sheet_decoy_peptides, self.decoy_peptides_header, self.decoy_peptides)

    def _parse_protein_details(self, sheet_protein_details):
        self._read_table(sheet_protein_details, self.protein_details_header, self.protein_details)

    def _parse_protein_summary(self, sheet_protein_summary):
        self._read_table(sheet_protein_summary, self.protein_summary_header, self.protein_summary)

    def _parse_quant_stats_info(self, sheet_quant_stats_info):
        #self._read_table_with_annotation(sheet_quant_stats_info, self.quant_stats_info_header, self.quant_stats_info)
        pass

    def _parse_mascat_quant_details(self, sheet_mascat_quant_details):
        self._read_table_with_annotation(sheet_mascat_quant_details, self.mascat_quant_details_header, self.mascat_quant_details)

    def _parse_libra_quant_details(self, sheet_libra_quant_details):
        self._read_table_with_annotation(sheet_libra_quant_details, self.libra_quant_details_header, self.libra_quant_details)

    def _parse_protein_expression(self, sheet_protein_expression):
        pass

    def _read_table(self, sheet, header, values):
        """
        :type sheet: openpyxl.worksheet.Worksheet
        :type header: list
        :type values: list
        """
        cols = sheet.get_highest_column()

        for i in range(cols):
            header.append(sheet.cell(row=0, column=i).value)

        # header omitted
        rows = sheet.get_highest_row() - 1
        for i in range(rows):
            values.append(map(lambda j: sheet.cell(row=i+1, column=j).value, range(cols)))

    def _read_table_with_annotation(self, sheet, header, values):
        """
        :type sheet: openpyxl.worksheet.Worksheet
        :type header: dict
        :type values: dict
        """
        cols = sheet.get_highest_column()

        current_group = "main"

        # Read header and group elements
        for i in range(cols):
            current_group = sheet.cell(row=0, column=i).value or current_group
            if not current_group in header:
                header[current_group] = []
            head = sheet.cell(row=1, column=i).value
            header[current_group].append(head)

        # header omitted, read values
        rows = sheet.get_highest_row() - 2
        for i in range(rows):
            col_offset = 0
            for head in header:
                if not head in values:
                    values[head] = []
                values[head].append(map(lambda j: sheet.cell(row=i+2, column=j+col_offset).value, range(len(header[head]))))
                col_offset += len(header[head])

    @commit_on_success
    def apply(self):
        self.detail.save()

        self.species.save(self.detail)

        if hasattr(self, "notify_progress"):
            outstr = "Assigning KEGG pathways"
            self.notify_progress(current=len(cds_map.values()), total=len(cds_map.values()), message=outstr)
